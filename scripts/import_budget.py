#!/usr/bin/env python3
"""
Импорт бюджетных данных из xlsx в PostgreSQL.
Парсит БДДС и БДР, определяет уровень статьи, сохраняет факт/план по месяцам.

Уровни:
  0 = ЦФО итого
  1 = БДДС/БДР итого
  2 = Раздел (- префикс)
  3 = Категория (plain)
  4 = Лист (* префикс)
"""

import sys
import os
import openpyxl
import psycopg2

DB_URL = "postgresql://indpark:indpark2026@127.0.0.1:5432/indparkdocs"
FILES = {
    'БДДС': '/tmp/budget_2026_bdds.xlsx',
    'БДР':  '/tmp/budget_2026_bdr.xlsx',
}

# Колонки: month_n_fact = 3 + n*4, month_n_plan = 4 + n*4 (n=0..11)
# total_fact = col[51], total_plan = col[52]

def get_level(name):
    if name.startswith('*'):
        return 4
    if name.startswith('-'):
        return 2
    return 3  # plain category

def parse_file(budget_type, filepath, target_cfos=None):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows_out = []
    cur_cfo = None

    for row in ws.iter_rows(min_row=6, values_only=True):
        name = row[0]
        if name is None:
            continue
        name = str(name).strip()
        if not name:
            continue

        # БДДС/БДР строка — итог ЦФО (уровень 1)
        if name in ('БДДС', 'БДР'):
            if cur_cfo is None:
                continue
            level = 1
        else:
            # Определяем: новый ЦФО или статья
            prev_was_total = len(rows_out) > 0 and rows_out[-1]['level'] in (0,)
            if not name.startswith('*') and not name.startswith('-') and \
               name not in ('ПОСТУПЛЕНИЯ', 'РАСХОДЫ', 'ДОХОДЫ',
                            'НАЛОГИ ИЗ ПРИБЫЛИ', 'Итого'):
                # Это может быть новый ЦФО (перед ним будет пустая строка или конец предыдущего)
                # ЦФО = строки без префикса, после которых идёт БДДС/БДР
                # Мы узнаём ЦФО ретроспективно — но можем проверить что следующая строка = БДДС/БДР
                # Вместо lookahead: ЦФО — строки с коротким именем (нет разделов)
                # Простой признак: если это строка-итог (уровень 0), то следующая = БДДС/БДР
                # Читаем 2 строки вперёд — не можем в read_only. 
                # Решение: первый проход собирает все строки, второй — определяет ЦФО.
                pass
            level = get_level(name)

        # Извлекаем данные
        fact = []
        plan = []
        for m in range(12):
            f = row[3 + m * 4]
            p = row[4 + m * 4]
            fact.append(float(f) if f is not None else None)
            plan.append(float(p) if p is not None else None)
        total_fact = row[51]
        total_plan = row[52]

        rows_out.append({
            'name': name,
            'fact': fact,
            'plan': plan,
            'total_fact': float(total_fact) if total_fact is not None else None,
            'total_plan': float(total_plan) if total_plan is not None else None,
        })

    wb.close()
    return rows_out


def parse_file_v2(budget_type, filepath, target_cfos=None):
    """
    Двухпроходный парсер:
    Проход 1: собираем все строки
    Проход 2: определяем ЦФО (строка перед БДДС/БДР)
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    raw = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        name = row[0]
        if name is None:
            continue
        name = str(name).strip()
        if not name:
            continue
        fact = [float(row[3 + m*4]) if row[3 + m*4] is not None else None for m in range(12)]
        plan = [float(row[4 + m*4]) if row[4 + m*4] is not None else None for m in range(12)]
        total_fact = float(row[51]) if row[51] is not None else None
        total_plan = float(row[52]) if row[52] is not None else None
        raw.append({'name': name, 'fact': fact, 'plan': plan, 'total_fact': total_fact, 'total_plan': total_plan})

    wb.close()

    # Определяем ЦФО: строка i является ЦФО если raw[i+1].name in ('БДДС', 'БДР')
    cfo_indices = set()
    for i, r in enumerate(raw):
        if i + 1 < len(raw) and raw[i+1]['name'] in ('БДДС', 'БДР'):
            cfo_indices.add(i)

    # Присваиваем ЦФО каждой строке
    cur_cfo = None
    result = []
    for i, r in enumerate(raw):
        name = r['name']
        if i in cfo_indices:
            cur_cfo = name
            level = 0
        elif name in ('БДДС', 'БДР'):
            level = 1
        elif name in ('ПОСТУПЛЕНИЯ', 'РАСХОДЫ', 'ДОХОДЫ', 'НАЛОГИ ИЗ ПРИБЫЛИ', 'Итого'):
            level = 2
        elif name.startswith('-'):
            level = 2
        elif name.startswith('*'):
            level = 4
        else:
            level = 3

        if cur_cfo is None:
            continue

        # Фильтр по ЦФО если задан
        if target_cfos and cur_cfo not in target_cfos:
            continue

        result.append({
            'budget_type': budget_type,
            'cfo': cur_cfo,
            'article': name,
            'level': level,
            'fact': r['fact'],
            'plan': r['plan'],
            'total_fact': r['total_fact'],
            'total_plan': r['total_plan'],
        })

    return result


def import_to_db(rows):
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()

    # Удаляем старые данные для тех же ЦФО
    cfos = list(set((r['budget_type'], r['cfo']) for r in rows))
    for bt, cfo in cfos:
        cur.execute("DELETE FROM budget_data WHERE budget_type=%s AND cfo=%s", (bt, cfo))

    inserted = 0
    for r in rows:
        cur.execute("""
            INSERT INTO budget_data (budget_type, cfo, article, level, fact, plan, total_fact, total_plan)
            VALUES (%s, %s, %s, %s, %s::numeric[], %s::numeric[], %s, %s)
            ON CONFLICT (budget_type, cfo, article) DO UPDATE
            SET level=EXCLUDED.level, fact=EXCLUDED.fact, plan=EXCLUDED.plan,
                total_fact=EXCLUDED.total_fact, total_plan=EXCLUDED.total_plan,
                imported_at=NOW()
        """, (
            r['budget_type'], r['cfo'], r['article'], r['level'],
            r['fact'], r['plan'], r['total_fact'], r['total_plan']
        ))
        inserted += 1

    conn.commit()
    cur.close()
    conn.close()
    return inserted


if __name__ == '__main__':
    # По умолчанию импортируем все ЦФО; можно передать список через аргументы
    target_cfos = sys.argv[1:] if len(sys.argv) > 1 else None
    if target_cfos:
        print(f"Импортируем ЦФО: {target_cfos}")
    else:
        print("Импортируем все ЦФО")

    all_rows = []
    for bt, fp in FILES.items():
        print(f"  Парсим {bt} ({fp})...")
        rows = parse_file_v2(bt, fp, target_cfos)
        print(f"    Строк: {len(rows)}")
        all_rows.extend(rows)

    print(f"Итого строк: {len(all_rows)}")
    print("Сохраняем в БД...")
    n = import_to_db(all_rows)
    print(f"Сохранено: {n} строк")

    # Покажем что получилось
    import psycopg2
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()
    cur.execute("SELECT budget_type, cfo, COUNT(*) FROM budget_data GROUP BY 1,2 ORDER BY 1,2")
    print("\nДанные в БД:")
    for row in cur.fetchall():
        print(f"  {row[0]:6} {row[1]:20} {row[2]:4} строк")
    cur.close()
    conn.close()
